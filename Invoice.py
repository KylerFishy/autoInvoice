import re
import xlrd
import xlwt
import os.path
import openpyxl
from datetime import datetime, timedelta
from time import sleep

# Used to add color/style to text (does not work in all OS system/terminals)
class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def lookForDate(line):
    if 'Date :' in line:
        date = line[9:19]
        try:
            day = getWeekdayFromDate(date)
        except:
            day = ""
        return date, day
    else:
        return False

def lookForEmailTime(line):
    if 'Date :' in line:
        return line[len(line)-9:].strip()
    else:
        return False

def lookForCallNum(line):
    if 'Cust. Service' in line:
        return line[26:].strip()
    else:
        return False

def lookForNameAndMerchantNum(line):
    if 'Service To' in line:
        name = line[14:len(line)-10].strip()
        merchantNum = line[len(line)-10:].strip()
        return name, merchantNum
    else:
        return False

def lookForTicket(line):
    ticketLine = re.findall(r'[/]\w+\s+\d{8}', line) # get ticket num
    if len(ticketLine): #if regex returned a result (length is not 0)
        return ticketLine[0][10:]
    else:
        return False

def lookForAddress(line, postalCodes):
    if 'Address' in line:
        postalCode = line[len(line)-7:].strip()
        address = line[13:].strip()
        if postalCode in postalCodes:
            km = int(postalCodes[postalCode][0])
            pay = '$' + str(postalCodes[postalCode][1])
            return postalCode, address, km, pay
        else:
            return postalCode, address, 0, 0
    else:
        return False

def getUserCommands():
    fileName = ""
    correctResponse = False
    while not correctResponse:
        usageResponse = input('Produce invoice data (i) or just print calls to console (c)? ')
        if usageResponse.lower() == 'i':
            fileName = input("Provide name for excel file (without extension): ")
            fileName += ".xlsx" # openpyxl library will only work with xlsx, not xls
            correctResponse = True
        elif usageResponse.lower() == 'c':
            correctResponse = True
        else:
            print('incorrect response, enter i or c')
            sleep(1)
    return usageResponse, fileName

def printCallSummary(call, usageResponse):
    print('\n\nName: ' + bcolors.BOLD + bcolors.FAIL + call['name'] + bcolors.ENDC)
    print('Date: ' + call['date'][:8] + bcolors.WARNING + call['date'][8:] + bcolors.ENDC)
    print('Address:', call['address'])
    print('Call issued at:', call['emailTime'])
    print('Ticket #:', call['ticketNum'])
    print('Call #:', call['callNum'])
    print('km:', call['km'])
    if usageResponse.lower() == 'c':
        print('Weekday: ' + bcolors.WARNING + call['day'] + bcolors.ENDC)

def getWeekdayFromDate(date):
    day = datetime.strptime(date, '%Y/%m/%d').weekday()
    if day == 0:
        return "Monday"
    elif day == 1:
        return "Tuesday"
    elif day == 2:
        return "Wednesday"
    elif day == 3:
        return "Thursday"
    elif day == 4:
        return "Friday"
    elif day == 5:
        return "Saturday"
    else:
        return "Sunday"

def promptForMissingFields(data):
    if data[0] == "":
        data[0] = input("Could not find date, provide if known: ")
    if data[1] == "":
        data[1] = input("Could not find call number, provide if known ")
    if data[2] == "":
        data[2] = input("Could not find ticket number, provide if known: ")
    if data[5] == "":
        data[5] = input("Could not find postal code, provide if known: ")
    if data[6] == "":
        data[6] = input("Could not find name, provide if known: ")
    if data[7] == "":
        data[7] = input("Could not find 'ins field', provide if known (ATB/RB): ")
    if data[9] == "":
        data[9] = input("Could not find call description, provide if known (service.. etc): ")
    if data[10] == "":
        data[10] = input("Could not find km, please provide: ")
    if data[12] == "":
        data[12] = input("Could not find pay, please provide: ")
        data[12] = float(data[12])
    return data

# called after writing an excel entry, clears all the data variables
def clearFields(data):
    return [''] * 13

# must have the postcodes file in directory (or else change the first func line)
# produces a dictionary that maps postal codes to km and price
def readPostalCodes():
    workbook = xlrd.open_workbook('postalCodes.xlsm')
    sheet = workbook.sheet_by_index(0)
    numCol = sheet.ncols
    numRows = sheet.nrows
    postalDictionary = {}

    for col in range(1, numRows):
        currentPostalCode = sheet.cell(col,3).value
        postalDictionary[currentPostalCode] = [sheet.cell(col,5).value, sheet.cell(col,6).value] #[km, price]
    return postalDictionary

# Inserts a new row to the excel file with the data from recent call item
def addEntryToExcel(fileName, data):

    if not os.path.exists(fileName):
        wb = openpyxl.Workbook()
        wb.save(fileName)

    wb = openpyxl.load_workbook(fileName)
    sheet = wb['Sheet']

    sheet.append(data)
    wb.save(fileName)

def excelEntryPrompt(call, fileName):
    workedIt = input("(q to quit)\nDid you work this " + bcolors.WARNING + call['day'] + bcolors.ENDC + " call (1/0)? \n\n")
    if workedIt.lower() == 'q':
        return False
    elif int(workedIt) == 1:
        call['time'] = input("Time on site (minutes)? ")
        call['time'] = str(timedelta(minutes=int(call['time']))) # format the time string
        call['time'] = call['time'][:len(call['time'])-3]

        call['numTerm']= input("Number of terminals? ")

        # Determine call type
        if call['callNum'] == "N/A":
            call['desc'] = input("No service call number found, provide call type/descrption: ")
        else:
            call['desc'] = "Service"

        # Format km
        if int(call['km']) <= 10 and int(call['km']) > 0:
            call['km'] = "0-10"
        elif int(km) <= 60:
            call['km'] = "11-60"

        # Determine 'ins' field
        if len(call['merchantNum']) == 9:
            if call['merchantNum'][0] == 'A': # if first character is A
                call['ins'] = 'ATB'
            else:
                call['ins'] = 'RB'
        else:
            ins = input("Could not determine 'ins' from merchant #, please provide (RB/ATB):")

        data = [call['date'], call['callNum'], call['ticketNum'], "", "", call['postalCode']
                , call['name'], call['ins'], call['time'], call['desc'], call['km'], call['numTerm'], call['pay']]
        promptForMissingFields(data)
        data[12] = data[12]+'0' # dirty fix to add extra zero to pay field
        addEntryToExcel(fileName, data)
        return True

def initializeCallObj():
    call = {}
    call['date'] = ""
    call['day'] = ""
    call['emailTime'] = ""
    call['name'] = ""
    call['merchantNum'] = ""
    call['postalCode'] = ""
    call['address'] = ""
    call['callNum'] = ""
    call['ticketNum'] = ""
    call['km'] = 0
    call['pay'] = 0
    return call
